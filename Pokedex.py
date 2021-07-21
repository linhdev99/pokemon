import openpyxl  # Connect the library
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Font, Fill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NORMAL=0
FIRE=1
WATER=2 
ELECTRIC=3
GRASS=4
ICE=5
FIGHTING=6
POISON=7
GROUND=8
FLYING=9
PSYCHIC=10
BUG=11
ROCK=12
GHOST=13
DRAGON=14
DARK=15
STEEL=16
FAIRY=17

NOEFF = 0.0
NTEFF = 0.25
NVEFF = 0.5
NREFF = 1.0
SUPEREFF = 2.0
SSEFF = 4.0

NAMETYPE = {
	NORMAL:["NORMAL","acac9b"],
	FIRE:["FIRE","fe4422"],
	WATER:["WATER","3399ff"],
	ELECTRIC:["ELECTRIC","ffcc33"],
	GRASS:["GRASS","77cc55"],
	ICE:["ICE","66ccff"],
	FIGHTING:["FIGHTING","bb5544"],
	POISON:["POISON","aa5599"],
	GROUND:["GROUND","ddbb55"],
	FLYING:["FLYING","8899ff"],
	PSYCHIC:["PSYCHIC","ff5599"],
	BUG:["BUG","aabb22"],
	ROCK:["ROCK","bbaa66"],
	GHOST:["GHOST","6666bb"],
	DRAGON:["DRAGON","7766ee"],
	DARK:["DARK","775544"],
	STEEL:["STEEL","a9a9ba"],
	FAIRY:["FAIRY","ee99ee"],
}
EFFECTTYPE = {
	NOEFF:["0","2e3436"],
	NVEFF:["1/2","a40000"],
	NREFF:["","ffffff"],
	SUPEREFF:["2","4e9a06"],
}
NAMEEFFECT = {
	NOEFF:"No Effect (0%)",
	NTEFF:"Not effective (25%)",
	NVEFF:"Not very effective (50%)",
	NREFF:"Normal (100%)",
	SUPEREFF:"Super Effective (200%)",
	SSEFF:"Super super effective (400%)",
}

class TypeTable:
	def __init__(self):
		self.table = list()
		self.length = 18
		self._initialTable()
  
	def _initialTable(self):
		temp = []
		for i in range(self.length):
			for j in range(self.length):
				temp.append(NREFF)
			self.table.append(temp)
			temp = []
		self._initEffDamage()

	def _effectDamage(self, _atk, _def, _eff):
		self.table[_atk][_def] = _eff
  
	def _initEffDamage(self):
		self._damageTypeNormal()
		self._damageTypeFire()
		self._damageTypeWater()
		self._damageTypeElectric()
		self._damageTypeGrass()
		self._damageTypeIce()
		self._damageTypeFighting()	
		self._damageTypePoison()
		self._damageTypeGround() 
		self._damageTypeFlying()
		self._damageTypePsychic()
		self._damageTypeBug()
		self._damageTypeRock()
		self._damageTypeGhost()
		self._damageTypeDragon()
		self._damageTypeDark()
		self._damageTypeSteel()
		self._damageTypeFairy()

	def _damageTypeNormal(self):
		_atk = NORMAL
		self._effectDamage(_atk, ROCK, NVEFF)
		self._effectDamage(_atk, GHOST, NOEFF)
		self._effectDamage(_atk, STEEL, NVEFF)
	
	def _damageTypeFire(self):
		_atk = FIRE
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, NVEFF)
		self._effectDamage(_atk, GRASS, SUPEREFF)
		self._effectDamage(_atk, ICE, SUPEREFF)
		self._effectDamage(_atk, BUG, SUPEREFF)
		self._effectDamage(_atk, ROCK, NVEFF)
		self._effectDamage(_atk, DRAGON, NVEFF)
		self._effectDamage(_atk, STEEL, SUPEREFF)
	
	def _damageTypeWater(self):
		_atk = WATER
		self._effectDamage(_atk, FIRE, SUPEREFF)
		self._effectDamage(_atk, WATER, NVEFF)
		self._effectDamage(_atk, GRASS, NVEFF)
		self._effectDamage(_atk, GROUND, SUPEREFF)
		self._effectDamage(_atk, ROCK, SUPEREFF)
		self._effectDamage(_atk, DRAGON, NVEFF)

	def _damageTypeElectric(self):
		_atk = ELECTRIC
		self._effectDamage(_atk, WATER, SUPEREFF)
		self._effectDamage(_atk, ELECTRIC, NVEFF)
		self._effectDamage(_atk, GRASS, NVEFF)
		self._effectDamage(_atk, GROUND, NOEFF)
		self._effectDamage(_atk, FLYING, SUPEREFF)
		self._effectDamage(_atk, DRAGON, NVEFF)
	
	def _damageTypeGrass(self):
		_atk = GRASS
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, SUPEREFF)
		self._effectDamage(_atk, GRASS, NVEFF)
		self._effectDamage(_atk, POISON, NVEFF)
		self._effectDamage(_atk, GROUND, SUPEREFF)
		self._effectDamage(_atk, FLYING, NVEFF)
		self._effectDamage(_atk, BUG, NVEFF)
		self._effectDamage(_atk, ROCK, SUPEREFF)
		self._effectDamage(_atk, DRAGON, NVEFF)
		self._effectDamage(_atk, STEEL, NVEFF)
  
	def _damageTypeIce(self):
		_atk = ICE
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, NVEFF)
		self._effectDamage(_atk, GRASS, SUPEREFF)
		self._effectDamage(_atk, ICE, NVEFF)
		self._effectDamage(_atk, GROUND, SUPEREFF)
		self._effectDamage(_atk, FLYING, SUPEREFF)
		self._effectDamage(_atk, DRAGON, SUPEREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
  
	def _damageTypeFighting(self):
		_atk = FIGHTING
		self._effectDamage(_atk, NORMAL, SUPEREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, SUPEREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, NVEFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NVEFF)
		self._effectDamage(_atk, PSYCHIC, NVEFF)
		self._effectDamage(_atk, BUG, NVEFF)
		self._effectDamage(_atk, ROCK, SUPEREFF)
		self._effectDamage(_atk, GHOST, NOEFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, SUPEREFF)
		self._effectDamage(_atk, STEEL, SUPEREFF)
		self._effectDamage(_atk, FAIRY, NVEFF)
				
	def _damageTypePoison(self):
		_atk = POISON
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, SUPEREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, NVEFF)
		self._effectDamage(_atk, GROUND, NVEFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NVEFF)
		self._effectDamage(_atk, GHOST, NVEFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, NOEFF)
		self._effectDamage(_atk, FAIRY, SUPEREFF)
     			
	def _damageTypeGround(self):
		_atk = GROUND
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, SUPEREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, SUPEREFF)
		self._effectDamage(_atk, GRASS, NVEFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, SUPEREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NOEFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, NVEFF)
		self._effectDamage(_atk, ROCK, SUPEREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, SUPEREFF)
		self._effectDamage(_atk, FAIRY, NREFF)
      			
	def _damageTypeFlying(self):
		_atk = FLYING
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NVEFF)
		self._effectDamage(_atk, GRASS, SUPEREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, SUPEREFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, SUPEREFF)
		self._effectDamage(_atk, ROCK, NVEFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NREFF)
    			
	def _damageTypePsychic(self):
		_atk = PSYCHIC
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, SUPEREFF)
		self._effectDamage(_atk, POISON, SUPEREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NVEFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NOEFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NREFF)
    			
	def _damageTypeBug(self):
		_atk = BUG
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, SUPEREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NVEFF)
		self._effectDamage(_atk, POISON, NVEFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NVEFF)
		self._effectDamage(_atk, PSYCHIC, SUPEREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, NVEFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, SUPEREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NVEFF)
    			
	def _damageTypeRock(self):
		_atk = ROCK
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, SUPEREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, SUPEREFF)
		self._effectDamage(_atk, FIGHTING, NVEFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NVEFF)
		self._effectDamage(_atk, FLYING, SUPEREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, SUPEREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NREFF)
      			
	def _damageTypeGhost(self):
		_atk = GHOST
		self._effectDamage(_atk, NORMAL, NOEFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, SUPEREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, SUPEREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NVEFF)
		self._effectDamage(_atk, STEEL, NREFF)
		self._effectDamage(_atk, FAIRY, NREFF)
    			
	def _damageTypeDragon(self):
		_atk = DRAGON
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, SUPEREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NOEFF)
    		
	def _damageTypeDark(self):
		_atk = DARK
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NREFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, NVEFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, SUPEREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, SUPEREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NVEFF)
		self._effectDamage(_atk, STEEL, NREFF)
		self._effectDamage(_atk, FAIRY, NVEFF)
    		
	def _damageTypeSteel(self):
		_atk = STEEL
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, NVEFF)
		self._effectDamage(_atk, ELECTRIC, NVEFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, SUPEREFF)
		self._effectDamage(_atk, FIGHTING, NREFF)
		self._effectDamage(_atk, POISON, NREFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, SUPEREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, NREFF)
		self._effectDamage(_atk, DARK, NREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, SUPEREFF)
    		
	def _damageTypeFairy(self):
		_atk = FAIRY
		self._effectDamage(_atk, NORMAL, NREFF)
		self._effectDamage(_atk, FIRE, NVEFF)
		self._effectDamage(_atk, WATER, NREFF)
		self._effectDamage(_atk, ELECTRIC, NREFF)
		self._effectDamage(_atk, GRASS, NREFF)
		self._effectDamage(_atk, ICE, NREFF)
		self._effectDamage(_atk, FIGHTING, SUPEREFF)
		self._effectDamage(_atk, POISON, NVEFF)
		self._effectDamage(_atk, GROUND, NREFF)
		self._effectDamage(_atk, FLYING, NREFF)
		self._effectDamage(_atk, PSYCHIC, NREFF)
		self._effectDamage(_atk, BUG, NREFF)
		self._effectDamage(_atk, ROCK, NREFF)
		self._effectDamage(_atk, GHOST, NREFF)
		self._effectDamage(_atk, DRAGON, SUPEREFF)
		self._effectDamage(_atk, DARK, SUPEREFF)
		self._effectDamage(_atk, STEEL, NVEFF)
		self._effectDamage(_atk, FAIRY, NREFF)
    
	def _exportXLSX(self):
		wb = Workbook()
		sheet_atk = wb.create_sheet(title='Attack',index=0)
		# sheet_atk = wb.active
		def setCellValue(_row, _col, _value, _size, _textColor='000000', _bgColor='FFFFFF'):
			temp = sheet_atk.cell(row=_row+1, column=_col+1)
			temp.fill = PatternFill(fill_type='solid', start_color=_bgColor, end_color=_bgColor)
			temp.font = Font(size=_size, color=_textColor, bold=True, shadow='000000', outline='000000')
			temp.alignment = Alignment(horizontal='center',vertical='center')
			sheet_atk.column_dimensions[get_column_letter(_col+1)].width = 10
			sheet_atk.row_dimensions[_row+1].height = 25
			temp.value = _value
			thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
			temp.border = thin_border
		i = 0
		j = 0
		setCellValue(i, j, "Atkꜜ|Def→", 10)
		j = 1
		for x in NAMETYPE:
			_type = NAMETYPE[x]
			setCellValue(i, j, _type[0], 12, 'FFFFFF', _type[1])
			j += 1
		i = 1
		j = 0
		for idx, x in enumerate(self.table):
			j = 0
			_type = NAMETYPE[idx]
			setCellValue(i, j, _type[0], 12, 'FFFFFF', _type[1])
			for y in x:
				j += 1 
				_type = EFFECTTYPE[y]
				setCellValue(i, j, _type[0], 12, 'f6dd3e', _type[1])
			i += 1
		wb.save('pokedex.xlsx')

class AttackType:
	def __init__(self):
		self.attack_table = TypeTable()
	def _checkAttack(self, _atk=None, _defs=None):
		if _atk != None and isinstance(_defs, list):
			value = 1.0
			for _def in _defs:
				value = value * self.attack_table.table[_atk][_def]
			return value 
		return None
	def _checkAll(self, _atk=None):
		if _atk != None:
			table = self.attack_table.table
			lst_noeff = []
			lst_nveff = []
			lst_nreff = []
			lst_sueff = []
			for x in NAMETYPE:
				# print(NAMETYPE[x][0], "=", table[_atk][x])
				eff = table[_atk][x]
				if eff == NOEFF:
					lst_noeff.append(NAMETYPE[x][0])
				elif eff == NVEFF:
					lst_nveff.append(NAMETYPE[x][0])
				elif eff == NREFF:
					lst_nreff.append(NAMETYPE[x][0])
				elif eff == SUPEREFF:
					lst_sueff.append(NAMETYPE[x][0])
			tab = {}
			tab[NOEFF] = lst_noeff
			tab[NVEFF] = lst_nveff
			tab[NREFF] = lst_nreff
			tab[SUPEREFF] = lst_sueff
			return tab
		return None

class DefenseType:
	def __init__(self):
		self.defense_table = TypeTable()
	def _checkDefense(self, _defs=None, _atk=None):
		if _atk != None and isinstance(_defs, list):
			value = 1.0
			for _def in _defs:
				value = value * self.defense_table.table[_atk][_def]
			return value
		return None
	def _checkAll(self, _defs=None):
		if _defs == None:
			return None
		if isinstance(_defs, list):
			table = self.defense_table.table
			lst_noeff = []
			lst_nteff = []
			lst_nveff = []
			lst_nreff = []
			lst_sueff = []
			lst_sseff = []
			for x in NAMETYPE:
				eff = self._checkDefense(_defs, x)
				if eff == NOEFF:
					lst_noeff.append(NAMETYPE[x][0])
				elif eff == NTEFF:
					lst_nteff.append(NAMETYPE[x][0])
				elif eff == NVEFF:
					lst_nveff.append(NAMETYPE[x][0])
				elif eff == NREFF:
					lst_nreff.append(NAMETYPE[x][0])
				elif eff == SUPEREFF:
					lst_sueff.append(NAMETYPE[x][0])
				elif eff == SSEFF:
					lst_sseff.append(NAMETYPE[x][0])
			tab = {}
			tab[NOEFF] = lst_noeff
			tab[NTEFF] = lst_nteff
			tab[NVEFF] = lst_nveff
			tab[NREFF] = lst_nreff
			tab[SUPEREFF] = lst_sueff
			tab[SSEFF] = lst_sseff
			return tab
		return None

def main():
	# attack_table = TypeTable()
	# attack_table._exportXLSX()
	attack = AttackType()
	print("GRASS(damage=100) attack (FIRE, FLYING) (damage=", 100 * attack._checkAttack(GRASS, [FIRE, FLYING]), ")")
	defense = DefenseType()
	print("(FIRE, FLYING) (recive damage=(", 100 * defense._checkDefense([FIRE], GRASS), ") from GRASS(100)")

	typePokemon = FIRE
	getData = attack._checkAll(typePokemon)
	print("=======\n", NAMETYPE[typePokemon][0], " TYPE ATTACK\n=======")
	for x in getData:
		print(NAMEEFFECT[x],":", getData[x])
	print("=======")

	typePokemon = [DRAGON, GRASS]
	getData = defense._checkAll(typePokemon)
	print("=======")
	print("[", NAMETYPE[typePokemon[0]][0], NAMETYPE[typePokemon[1]][0], "] TYPE DEFENSE")
	print("=======")
	for y in getData:
		print(NAMEEFFECT[y],":", getData[y])
	print("=======")

	# print(defense._checkAll([FIRE]))
 
	# getData = attack._checkAll(FIRE)
	# print("=======\nFIRE TYPE\n=======")
	# for x in getData:
	# 	print(NAMEEFFECT[x],":", getData[x])
	# print("=======")
	# for x in NAMETYPE:
	# 	getData = attack._checkAll(x)
	# 	print("=======")
	# 	print(NAMETYPE[x][0], "TYPE")
	# 	print("=======")
	# 	for y in getData:
	# 		print(NAMEEFFECT[y],":", getData[y])
	# 	print("=======")
	# for x in NAMETYPE:
	# 	getData = defense._checkAll([x])
	# 	print("=======")
	# 	print(NAMETYPE[x][0], "TYPE DEFENSE")
	# 	print("=======")
	# 	for y in getData:
	# 		print(NAMEEFFECT[y],":", getData[y])
	# 	print("=======")

if __name__ == '__main__':
	main()
